import tkinter as tk
from openpyxl import Workbook

gastos = []

def agregar_gasto():
    monto = entrada_monto.get()
    descripcion = entrada_descripcion.get()

    if monto != "" and descripcion != "":
        try:
            monto = float(monto)

            gastos.append((descripcion, monto))

            lista.insert(
                tk.END,
                f"{descripcion} - $ {monto:.2f}"
            )

            entrada_monto.delete(0, tk.END)
            entrada_descripcion.delete(0, tk.END)

            actualizar_total()

        except ValueError:
            etiqueta_total.config(
                text="Ingrese un importe válido"
            )

def actualizar_total():
    total = sum(monto for descripcion, monto in gastos)

    etiqueta_total.config(
        text=f"Total: $ {total:.2f}"
    )

def guardar_excel():

    wb = Workbook()
    ws = wb.active

    ws.title = "Gastos"

    ws["A1"] = "Descripción"
    ws["B1"] = "Importe"

    fila = 2

    for descripcion, monto in gastos:
        ws.cell(row=fila, column=1, value=descripcion)
        ws.cell(row=fila, column=2, value=monto)
        fila += 1

    ws.cell(row=fila + 1, column=1, value="TOTAL")
    ws.cell(row=fila + 1, column=2, value=sum(monto for _, monto in gastos))

    wb.save("gastos.xlsx")

    etiqueta_total.config(
        text="Archivo Excel guardado correctamente"
    )

# Ventana
ventana = tk.Tk()
ventana.title("Control de Gastos")
ventana.geometry("450x450")

# Descripción
tk.Label(
    ventana,
    text="Descripción"
).pack()

entrada_descripcion = tk.Entry(
    ventana,
    width=40
)
entrada_descripcion.pack(pady=5)

# Monto
tk.Label(
    ventana,
    text="Monto"
).pack()

entrada_monto = tk.Entry(
    ventana,
    width=20
)
entrada_monto.pack(pady=5)

# Botón agregar
boton_agregar = tk.Button(
    ventana,
    text="Agregar gasto",
    command=agregar_gasto
)
boton_agregar.pack(pady=10)

# Lista
lista = tk.Listbox(
    ventana,
    width=50,
    height=10
)
lista.pack(pady=10)

# Total
etiqueta_total = tk.Label(
    ventana,
    text="Total: $ 0",
    font=("Arial", 12, "bold")
)
etiqueta_total.pack(pady=10)

# Guardar Excel
boton_guardar = tk.Button(
    ventana,
    text="Guardar en Excel",
    command=guardar_excel
)
boton_guardar.pack(pady=10)

ventana.mainloop()